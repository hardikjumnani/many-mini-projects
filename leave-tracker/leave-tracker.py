# openpyxl – To update and read from the excel sheet.
# smtplib & email – To send out emails.

import smtplib
from email.message import EmailMessage

from typing import List, Dict
from openpyxl import load_workbook

from password import PASSWORD

SUB_NAME = 'Discrete Mathematics'
SUB_TEACHER_EMAIL = 'hardik.jumnani123@gmail.com'
TOTAL_CLASSES = 20
MIN_ATTENDANCE_PERCENT = 80

STU_ROLL_NO_COL_POS = 1
STU_NAME_COL_POS = 2
STU_EMAIL_COL_POS = 3

ATTENDANCE_START_COL_POS = 4 # 1-indexed

try: workbook = load_workbook('attendance_records.xlsx', data_only=True) # data_only to get vals of formulas
except:
    print('Create a new excel attendance sheet and name it attendance_records.xlsx')
    quit()

sheet = workbook['Sheet1']

mailing_list: Dict[str, List[str]] = {'one_leave_posb': [], 'no_leave_posb': []}

# Analyze mailing list
rows = list(sheet.iter_rows(min_row=1, values_only=True))
for i in range(1, len(rows)):
    row = rows[i]
    present: int = 0
    for i in range(ATTENDANCE_START_COL_POS-1, len(row)):
        if int(row[i]) > 0: present += 1

    one_absent_fine: bool = (present-1) / TOTAL_CLASSES >= MIN_ATTENDANCE_PERCENT/100
    two_absent_issue: bool = (present-2) / TOTAL_CLASSES < MIN_ATTENDANCE_PERCENT/100
    curr_attendance_fine: bool = (present) / TOTAL_CLASSES >= MIN_ATTENDANCE_PERCENT/100
    no_absent_posb: bool = (present-1) / TOTAL_CLASSES < MIN_ATTENDANCE_PERCENT/100
    
    one_leave_posb: bool = one_absent_fine and two_absent_issue
    no_leave_posb: bool = curr_attendance_fine and no_absent_posb

    if one_leave_posb:
        mailing_list['one_leave_posb'].append(row)
    elif no_leave_posb:
        mailing_list['no_leave_posb'].append(row)

# Send mails

# Email Configuration
smtp_server = "smtp.gmail.com"
port = 587
sender_email = "hardik.jumnani123@gmail.com"
password = PASSWORD
# receiver_email = "receiver_email@gmail.com"

emails_to_send = []

for student in mailing_list['one_leave_posb']:
    msg = EmailMessage()
    msg["From"] = sender_email
    msg["To"] = student[STU_EMAIL_COL_POS-1]
    msg["Subject"] = "URGENT: 1 Leave Left, Reminder for" + SUB_NAME

    msg.set_content('''Student,\nthis is to remind you that you only have \
1 leave left for the subject. Illness is inevitable but \
take care of yourself and avoid missing anymore classes as \
it will affect your grades.''')

    emails_to_send.append(msg)

teacher_mail_content = f'''This is to inform you that the following students don't have\
any leaves left for the subject, {SUB_NAME}. Kindly \
make sure the student takes no more leaves.\n'''

for student in mailing_list['no_leave_posb']:
    msg = EmailMessage()
    msg["From"] = sender_email
    msg["To"] = student[STU_EMAIL_COL_POS-1]
    msg["Subject"] = "URGENT: No Leave Left, Reminder for" + SUB_NAME

    msg.set_content('''Student,\nthis is to remind you that you have exhausted \
all your leaves for this course. Taking another leave \
shall reduce your grades achieved by the course-end exams.\
Kindly ensure this isn't the case.''')
    emails_to_send.append(msg)

    teacher_mail_content += f"{student[STU_ROLL_NO_COL_POS-1]}\t{student[STU_NAME_COL_POS-1]}\t{student[STU_EMAIL_COL_POS-1]}\n"

# mail to teacher with list of students
msg = EmailMessage()
msg["From"] = sender_email
msg["To"] = SUB_TEACHER_EMAIL
msg["Subject"] = f"Student who have no leaves left."

msg.set_content(teacher_mail_content)

emails_to_send.append(msg)

print([i['To'] for i in emails_to_send])

# send mails
try:
    with smtplib.SMTP(smtp_server, port) as server:
        server.starttls()  # Secure the connection
        server.login(sender_email, password)

        for msg in emails_to_send:
            server.send_message(msg)
            print(f"Email sent to {msg['To']}!")

except Exception as e:
    print(f"Failed to send email: {e}")